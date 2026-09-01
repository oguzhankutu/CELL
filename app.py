import streamlit as st
import pandas as pd
import openpyxl
import re
import datetime
import io

# Web sayfası tasarımı
st.set_page_config(page_title="Satış Raporu Otomasyonu", page_icon="📈", layout="wide")
st.title("📈 Kuzeyboru Satış Raporu Dönüştürücü")
st.markdown("Sipariş raporunuzu yükleyin; **Ezgi Genç** ve **Gökhan Çelebi** için haftalık ve aylık özel özet tablonuzu anında indirin.")

# Dosya Yükleme Kutusu
yuklenen_dosya = st.file_uploader("Excel Dosyasını Sürükleyin (.xls veya .xlsx)", type=["xls", "xlsx"])

if yuklenen_dosya is not None:
    if st.button("🚀 Raporu Hazırla ve İndir"):
        with st.spinner("Yapay zeka verilerinizi analiz ediyor, tabloyu şekillendiriyor..."):
            try:
                # 1. DOSYA OKUMA VE BAŞLIK TESPİTİ
                try:
                    df = pd.read_excel(yuklenen_dosya, header=1)
                    if 'TARİH' not in df.columns and 'SİPARİŞ_NO' not in df.columns:
                        df = pd.read_excel(yuklenen_dosya, header=0)
                except Exception:
                    df = pd.read_excel(yuklenen_dosya, header=0)

                df = df.dropna(how='all').reset_index(drop=True)

                # 2. SÜTUN İSİMLERİNİ TEMİZLEME
                def standardize_column_name(col_name):
                    col_name = str(col_name).strip().replace('\n', '').replace('\r', '')
                    karakter_haritasi = str.maketrans('İŞĞÜÖÇışğüöç', 'ISGUOCisguoc')
                    return col_name.translate(karakter_haritasi).lower()

                def turkish_upper(text):
                    if not isinstance(text, str): return str(text)
                    mapping = {'i': 'İ', 'ı': 'I', 'ğ': 'Ğ', 'ü': 'Ü', 'ş': 'Ş', 'ö': 'Ö', 'ç': 'Ç'}
                    for k, v in mapping.items(): text = text.replace(k, v)
                    return text.upper().strip()

                df.rename(columns={col: standardize_column_name(col) for col in df.columns}, inplace=True)

                sutun_haritasi = {}
                for col in df.columns:
                    if 'temsilci' in col or 'satici' in col: sutun_haritasi[col] = 'SATICI'
                    elif 'alt' in col and 'urun' in col: sutun_haritasi[col] = 'ALT_URUN_GRUBU'
                    elif 'tarih' in col: sutun_haritasi[col] = 'TARIH'
                    elif 'kg' in col and 'miktar' in col and 'sevk' not in col: sutun_haritasi[col] = 'KG_MIKTAR'
                    elif 'tl' in col and 'tutar' in col and 'bekleyen' not in col: sutun_haritasi[col] = 'TL_TUTAR'

                df.rename(columns=sutun_haritasi, inplace=True)

                # 3. KİŞİ FİLTRELEME
                istenilen_temsilciler = ['EZGİ GENÇ', 'GÖKHAN ÇELEBİ']
                df['SATICI'] = df['SATICI'].apply(turkish_upper)
                df = df[df['SATICI'].isin(istenilen_temsilciler)].copy()

                # 4. KUSURSUZ GRUPLANDIRMA (Yeni Kurallarınız)
                grup_tanimlari = {
                    'HDPE': ['HDPE', 'HDPE BORU', 'HDPE DRENAJ BORU', 'HDPE DRENAJ BORUSU', 'HDPE DRENAJLI BORU', 'LDPE', 'LDPE BORU'],
                    'KORUGE': ['ÇELİK TAKVİYELİ KORUGE BORU', 'KORUGE DRENAJ BORU', 'KORUGE BORU'],
                    'PPR': ['PPR', 'PPR BORU', 'PP-R BORU', 'PVC', 'ATIK SU PVC BORU', 'PE-RT BORU', 'PVC BORU'],
                    'EK PARÇA': ['ÇELİK FLANŞ', 'EF KAYNAK', 'EF KAYNAK MAKİNESİ', 'ALIN KAYNAK MAKİNESİ', 'EK PARÇA', 'DİĞER EK PARÇA', 'KABLO MUHAFAZA KORUGE BORU', 'KABLO MUHAFAZALI KORUGE BORU', 'KORUGE GEOTEKSTİL DRENAJ BORU', 'KORUGE GEOTEKSTİL DRENAJ BORUSU', 'PPR EK PARÇA', 'PP-R EK PARÇA', 'PVC EK PARÇA', 'ATIK SU PVC EK PARÇA', 'HDPE EK PARÇA', 'KORUGE EK PARÇA'],
                    'ÜST YAPI': ['METAL ÜRÜNLER', 'PE-RT', 'PE-RT - PE-XB BORU', 'PE-RT -PE-XB BORU', 'PPR EK PARÇA', 'PP-R EK PARÇA', 'PVC EK PARÇA']
                }

                grup_tanimlari_buyuk = {ana_grup: [turkish_upper(u) for u in alt_urunler] for ana_grup, alt_urunler in grup_tanimlari.items()}

                def grubu_bul(urun_adi):
                    urun_adi = turkish_upper(urun_adi)
                    if urun_adi in ('NAN', '', 'NONE', 'HİZMET'): return 'HİZMET'
                    for ana_grup, alt_urunler_buyuk in grup_tanimlari_buyuk.items():
                        if urun_adi in alt_urunler_buyuk: return ana_grup
                    return 'DİĞER'

                df['ALT_URUN_GRUBU'] = df['ALT_URUN_GRUBU'].fillna('')
                df['GRUP'] = df['ALT_URUN_GRUBU'].apply(grubu_bul)

                # 5. MATEMATİKSEL TEMİZLİK (Tonaj ve TL Hesaplama)
                def robust_numeric(series):
                    def convert_val(x):
                        if pd.isna(x): return 0.0
                        if isinstance(x, (int, float)): return float(x)
                        x = str(x).strip()
                        if not x: return 0.0
                        if '.' in x and ',' in x:
                            if x.rfind('.') > x.rfind(','): x = x.replace(',', '')
                            else: x = x.replace('.', '').replace(',', '.')
                        elif ',' in x: x = x.replace(',', '.')
                        elif x.count('.') > 1: x = x.replace('.', '')
                        x = re.sub(r'[^\d.-]', '', x)
                        try: return float(x)
                        except ValueError: return 0.0
                    return series.apply(convert_val)

                df['TON'] = robust_numeric(df['KG_MIKTAR']) / 1000.0
                df['TL_TUTAR'] = robust_numeric(df['TL_TUTAR'])

                # 6. ZAMAN AYRIŞTIRMA
                df['TARIH'] = pd.to_datetime(df['TARIH'], errors='coerce', dayfirst=True)
                df['AY'] = df['TARIH'].dt.month
                df['HAFTA'] = df['TARIH'].dt.isocalendar().week

                month_map = {1: 'Ocak', 2: 'Şubat', 3: 'Mart', 4: 'Nisan', 5: 'Mayıs', 6: 'Haziran',
                             7: 'Temmuz', 8: 'Ağustos', 9: 'Eylül', 10: 'Ekim', 11: 'Kasım', 12: 'Aralık'}

                df['AY_AD'] = df['AY'].map(month_map).fillna('Tarihsiz')
                months_order = ['Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran', 'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık', 'Tarihsiz']

                df_rapor = df[df['GRUP'] != 'DİĞER'].copy()
                tum_gruplar = list(grup_tanimlari.keys())
                tam_index = pd.MultiIndex.from_product([istenilen_temsilciler, tum_gruplar], names=['SATICI', 'GRUP'])

                # 7. PİVOT TABLOLAR
                aylik_pivot = df_rapor.pivot_table(index=['SATICI', 'GRUP'], columns='AY_AD', values=['TON', 'TL_TUTAR'], aggfunc='sum', fill_value=0)
                aylik_pivot = aylik_pivot.reindex(tam_index, fill_value=0)
                aylik_pivot.columns = aylik_pivot.columns.swaplevel(0, 1)
                mevcut_aylar = [ay for ay in months_order if ay in aylik_pivot.columns.levels[0]]
                aylik_pivot = aylik_pivot.reindex(columns=mevcut_aylar, level=0)
                aylik_pivot = aylik_pivot.reindex(columns=['TON', 'TL_TUTAR'], level=1)
                
                aylik_pivot.loc[('Genel Toplam', ''), :] = aylik_pivot.sum(axis=0)
                aylik_pivot[('Genel Toplam', 'TON')] = aylik_pivot.xs('TON', level=1, axis=1).sum(axis=1)
                aylik_pivot[('Genel Toplam', 'TL_TUTAR')] = aylik_pivot.xs('TL_TUTAR', level=1, axis=1).sum(axis=1)

                haftalik_pivot = df_rapor.pivot_table(index=['SATICI', 'GRUP'], columns='HAFTA', values=['TON', 'TL_TUTAR'], aggfunc='sum', fill_value=0)
                haftalik_pivot = haftalik_pivot.reindex(tam_index, fill_value=0)
                haftalik_pivot.columns = haftalik_pivot.columns.swaplevel(0, 1)
                mevcut_haftalar = sorted([h for h in haftalik_pivot.columns.levels[0] if pd.notna(h)])
                haftalik_pivot = haftalik_pivot.reindex(columns=mevcut_haftalar, level=0)
                haftalik_pivot = haftalik_pivot.reindex(columns=['TON', 'TL_TUTAR'], level=1)
                haftalik_pivot.columns = haftalik_pivot.columns.set_levels([f"{int(w)}. Hafta" for w in haftalik_pivot.columns.levels[0]], level=0)
                
                haftalik_pivot.loc[('Genel Toplam', ''), :] = haftalik_pivot.sum(axis=0)
                haftalik_pivot[('Genel Toplam', 'TON')] = haftalik_pivot.xs('TON', level=1, axis=1).sum(axis=1)
                haftalik_pivot[('Genel Toplam', 'TL_TUTAR')] = haftalik_pivot.xs('TL_TUTAR', level=1, axis=1).sum(axis=1)

                # 8. BELLEKTE EXCEL FORMATLAMA (6. SÜTUNDAN BAŞLATMA VE BOŞLUKLAR)
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    haftalik_pivot.to_excel(writer, sheet_name='Haftalik_Ozet', merge_cells=True, startcol=5)
                    aylik_pivot.to_excel(writer, sheet_name='Aylik_Ozet', merge_cells=True, startcol=5)

                output.seek(0)
                wb = openpyxl.load_workbook(output)

                for sheet_name in ['Haftalik_Ozet', 'Aylik_Ozet']:
                    ws = wb[sheet_name]
                    
                    merged_ranges = list(ws.merged_cells.ranges)
                    for mr in merged_ranges:
                        if mr.min_col == 6 and mr.min_row >= 4:
                            top_val = ws.cell(row=mr.min_row, column=6).value
                            ws.unmerge_cells(str(mr))
                            for r in range(mr.min_row, mr.max_row + 1):
                                ws.cell(row=r, column=6).value = top_val

                    aktif_isim = None
                    for i in range(4, ws.max_row + 1):
                        hucre = ws.cell(row=i, column=6).value
                        if hucre is not None and str(hucre).strip() != '' and hucre != 'Genel Toplam':
                            aktif_isim = hucre
                        elif hucre is None or str(hucre).strip() == '':
                            if aktif_isim is not None:
                                ws.cell(row=i, column=6).value = aktif_isim

                    for i in range(ws.max_row, 4, -1):
                        val = ws.cell(row=i, column=6).value
                        prev_val = ws.cell(row=i-1, column=6).value
                        if val != prev_val and prev_val not in [None, 'SATICI']:
                            ws.insert_rows(i, amount=3)

                final_excel = io.BytesIO()
                wb.save(final_excel)
                final_excel.seek(0)

                st.success("✅ Veriler işlendi, hatalar temizlendi ve Excel'iniz hazır!")
                
                st.download_button(
                    label="📥 Kusursuz Excel Raporunu İndir",
                    data=final_excel,
                    file_name="Guncel_Haftalik_Aylik_Rapor.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"Dosya analiz edilirken bir hata oluştu: {e}")
